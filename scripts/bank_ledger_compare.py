#!/usr/bin/env python3
"""Compare QuickBooks bank/cash COA balances and ledgers with ERP (company 3)."""
from __future__ import annotations

import argparse
import csv
import re
import subprocess
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

COMPANY_ID = 3
TOLERANCE = Decimal("0.02")
SCRIPT_DIR = Path(__file__).resolve().parent
TB_CHECK = SCRIPT_DIR / "tb_check.csv"

# QuickBooks account number -> ERP account number (MIA company)
QB_TO_ERP = {
    "10800": "10015",
    "10900": "10016",
    "12000": "10017",
    "10020": "10013",
    "11000": "11110",
    "15200": "15100",
    "30800": "30020",
    "32000": "30000",
}

ERP_BANK_ACCOUNTS = [
    "10001", "10002", "10003", "10004", "10005", "10006", "10007", "10008",
    "10009", "10010", "10011", "10012", "10013", "10014", "10015", "10016", "10017",
]


def parse_decimal(value: str | None) -> Decimal:
    if not value or not str(value).strip():
        return Decimal("0")
    s = (
        str(value)
        .strip()
        .strip('"')
        .replace(",", "")
        .replace("(", "-")
        .replace(")", "")
    )
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def map_qb_account(number: str) -> str:
    return QB_TO_ERP.get(number.strip(), number.strip())


def load_qb_trial_balance() -> dict[str, Decimal]:
    balances: dict[str, Decimal] = {}
    if not TB_CHECK.exists():
        return balances
    with TB_CHECK.open(encoding="utf-8") as handle:
        for row in csv.DictReader(handle):
            acct = row.get("AccountNumber", "").strip()
            if acct in ERP_BANK_ACCOUNTS:
                balances[acct] = parse_decimal(row.get("net"))
    return balances


def run_sql(query: str) -> str:
    result = subprocess.run(
        ["sqlcmd", "-S", "localhost", "-d", "PakistanAccountingERP", "-E", "-W", "-s", "|"],
        input=query,
        capture_output=True,
        text=True,
        check=True,
    )
    return result.stdout


def load_erp_bank_closings() -> dict[str, Decimal]:
    sql = f"""
SET NOCOUNT ON;
SELECT
    coa.AccountNumber,
    coa.AccountName,
    coa.OpeningBalance,
    ISNULL(jt.Dr, 0) AS TotalDebit,
    ISNULL(jt.Cr, 0) AS TotalCredit,
    coa.OpeningBalance + ISNULL(jt.Dr, 0) - ISNULL(jt.Cr, 0) AS Closing
FROM ChartOfAccounts coa
CROSS APPLY (
    SELECT
        SUM(jel.Debit) AS Dr,
        SUM(jel.Credit) AS Cr
    FROM JournalEntryLines jel
    INNER JOIN JournalEntries je ON je.Id = jel.JournalEntryId
    WHERE jel.ChartOfAccountId = coa.Id
      AND je.CompanyId = {COMPANY_ID}
      AND je.IsDeleted = 0
      AND je.Status = 2
) jt
WHERE coa.CompanyId = {COMPANY_ID}
  AND coa.IsDeleted = 0
  AND coa.SubTypeId = 1
  AND NOT EXISTS (
      SELECT 1
      FROM ChartOfAccounts c2
      WHERE c2.ParentAccountId = coa.Id AND c2.IsDeleted = 0
  )
ORDER BY coa.AccountNumber;
"""
    closings: dict[str, Decimal] = {}
    for line in run_sql(sql).splitlines():
        if "|" not in line or line.startswith("AccountNumber") or line.startswith("-"):
            continue
        parts = [part.strip() for part in line.split("|")]
        if len(parts) < 6 or not parts[0].startswith("100"):
            continue
        closings[parts[0]] = parse_decimal(parts[5])
    return closings


def load_erp_ledger(account_number: str) -> list[dict]:
    sql = f"""
SET NOCOUNT ON;
SELECT
    je.EntryDate,
    je.EntryNumber,
    je.ReferenceType,
    ISNULL(jel.Memo, je.Description) AS LineMemo,
    jel.Debit,
    jel.Credit
FROM ChartOfAccounts coa
INNER JOIN JournalEntryLines jel ON jel.ChartOfAccountId = coa.Id
INNER JOIN JournalEntries je ON je.Id = jel.JournalEntryId
WHERE coa.CompanyId = {COMPANY_ID}
  AND coa.AccountNumber = N'{account_number}'
  AND coa.IsDeleted = 0
  AND je.CompanyId = {COMPANY_ID}
  AND je.IsDeleted = 0
  AND je.Status = 2
ORDER BY je.EntryDate, je.Id, jel.Id;
"""
    rows: list[dict] = []
    for line in run_sql(sql).splitlines():
        if "|" not in line or line.startswith("EntryDate") or line.startswith("-"):
            continue
        parts = [part.strip() for part in line.split("|")]
        if len(parts) < 6:
            continue
        rows.append(
            {
                "date": parts[0][:10],
                "entry": parts[1],
                "ref_type": parts[2],
                "memo": parts[3],
                "debit": parse_decimal(parts[4]),
                "credit": parse_decimal(parts[5]),
            }
        )
    return rows


def read_qb_excel_ledger(path: Path) -> tuple[Decimal | None, list[dict], Decimal | None]:
    """Read a QuickBooks Desktop General Ledger Excel export for one account."""
    if load_workbook is None:
        raise RuntimeError("openpyxl is required: pip install openpyxl")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "Sheet1" if "Sheet1" in workbook.sheetnames else workbook.sheetnames[-1]
    sheet = workbook[sheet_name]

    opening: Decimal | None = None
    closing: Decimal | None = None
    rows: list[dict] = []

    for row in sheet.iter_rows(min_row=1, values_only=True):
        cells = ["" if cell is None else str(cell).strip() for cell in row]
        if not any(cells):
            continue

        # QuickBooks Desktop export: columns at 0,5,7,9,11,13,15,17,19 (sparse layout)
        label = str(cells[1] or cells[2] or "").strip()
        if "10008" in label and opening is None and len(cells) > 19 and cells[19] is not None and not cells[5]:
            opening = parse_decimal(cells[19])
            continue
        if "Total 10008" in label:
            closing = parse_decimal(cells[19] if len(cells) > 19 else "0")
            continue

        txn_type = cells[5] if len(cells) > 5 else None
        txn_date = cells[7] if len(cells) > 7 else None
        if txn_type is None or txn_date is None:
            continue

        debit = parse_decimal(cells[15] if len(cells) > 15 else "0")
        credit = parse_decimal(cells[17] if len(cells) > 17 else "0")
        if debit == 0 and credit == 0:
            continue

        date_str = (
            txn_date.strftime("%Y-%m-%d")
            if hasattr(txn_date, "strftime")
            else str(txn_date)[:10]
        )
        rows.append(
            {
                "date": date_str,
                "type": str(txn_type).strip(),
                "num": str(cells[9] if len(cells) > 9 and cells[9] else "").strip(),
                "name": str(cells[11] if len(cells) > 11 and cells[11] else "").strip(),
                "memo": str(cells[13] if len(cells) > 13 and cells[13] else "").strip(),
                "debit": debit,
                "credit": credit,
                "balance": parse_decimal(cells[19] if len(cells) > 19 else "0") or None,
            }
        )

    workbook.close()
    if closing is None and opening is not None:
        net = sum(r["debit"] - r["credit"] for r in rows)
        closing = (opening + net).quantize(Decimal("0.01"), ROUND_HALF_UP)
    return opening, rows, closing


def compare_tb_vs_erp() -> int:
    qb = load_qb_trial_balance()
    erp = load_erp_bank_closings()

    print("=" * 88)
    print("BANK / CASH ACCOUNTS — QuickBooks TB (31-May-2026 cutover) vs ERP GL closing (now)")
    print("=" * 88)
    print(f"{'Acct':<8} {'QB TB':>18} {'ERP Closing':>18} {'Diff':>18}  Note")
    print("-" * 88)

    mismatches = 0
    for acct in ERP_BANK_ACCOUNTS:
        qb_val = qb.get(acct)
        erp_val = erp.get(acct)
        if qb_val is None and erp_val is None:
            continue
        diff = (erp_val or Decimal("0")) - (qb_val or Decimal("0"))
        if erp_val is None:
            note = "missing in ERP"
            mismatches += 1
        elif qb_val is None:
            note = "no QB TB row"
        elif abs(diff) <= TOLERANCE:
            note = "static (no post-cutover GL)" if diff == 0 else "June activity"
            if abs(diff) > TOLERANCE:
                mismatches += 1
                note = "MISMATCH"
        else:
            mismatches += 1
            note = "MISMATCH — check ledger"

        qb_s = f"{qb_val:,.2f}" if qb_val is not None else "N/A"
        erp_s = f"{erp_val:,.2f}" if erp_val is not None else "N/A"
        print(f"{acct:<8} {qb_s:>18} {erp_s:>18} {diff:>18,.2f}  {note}")

    print()
    print(f"Accounts checked: {len(ERP_BANK_ACCOUNTS)} | Mismatches vs QB TB: {mismatches}")
    return mismatches


def compare_excel_ledger(excel_path: Path, erp_account: str) -> int:
    qb_open, qb_rows, qb_close = read_qb_excel_ledger(excel_path)
    erp_rows = load_erp_ledger(erp_account)
    erp_open = parse_decimal(
        next(
            (
                line.split("|")[2].strip()
                for line in run_sql(
                    f"SET NOCOUNT ON; SELECT OpeningBalance FROM ChartOfAccounts "
                    f"WHERE CompanyId={COMPANY_ID} AND AccountNumber=N'{erp_account}' AND IsDeleted=0;"
                ).splitlines()
                if "|" in line and not line.startswith("-") and "OpeningBalance" not in line
            ),
            "0",
        )
    )
    erp_close = erp_open + sum(r["debit"] - r["credit"] for r in erp_rows)

    print("=" * 88)
    print(f"LEDGER COMPARE — QB Excel: {excel_path.name}  ->  ERP {erp_account}")
    print("=" * 88)
    print(f"QB opening:  {qb_open:,.2f}" if qb_open is not None else "QB opening:  (not found)")
    print(f"ERP opening: {erp_open:,.2f}")
    print(f"QB lines: {len(qb_rows)} | ERP lines: {len(erp_rows)}")
    print(f"QB closing:  {qb_close:,.2f}" if qb_close is not None else "QB closing:  (not found)")
    print(f"ERP closing: {erp_close:,.2f}")
    if qb_close is not None and abs(qb_close - erp_close) > TOLERANCE:
        print(f"CLOSING GAP: {(erp_close - qb_close):,.2f}")
    print()

    qb_debits = sum(r["debit"] for r in qb_rows)
    qb_credits = sum(r["credit"] for r in qb_rows)
    erp_debits = sum(r["debit"] for r in erp_rows)
    erp_credits = sum(r["credit"] for r in erp_rows)
    print(f"QB movement   Dr {qb_debits:,.2f}  Cr {qb_credits:,.2f}")
    print(f"ERP movement  Dr {erp_debits:,.2f}  Cr {erp_credits:,.2f}")
    print(f"Debit diff:  {(erp_debits - qb_debits):,.2f}")
    print(f"Credit diff: {(erp_credits - qb_credits):,.2f}")
    return 0 if qb_close is None or abs(qb_close - erp_close) <= TOLERANCE else 1


def export_erp_ledgers(output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for acct in ERP_BANK_ACCOUNTS:
        rows = load_erp_ledger(acct)
        if not rows:
            continue
        out = output_dir / f"erp_ledger_{acct}.csv"
        with out.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(
                handle,
                fieldnames=["date", "entry", "ref_type", "memo", "debit", "credit"],
            )
            writer.writeheader()
            writer.writerows(rows)
        print(f"Wrote {out} ({len(rows)} lines)")


def main() -> int:
    parser = argparse.ArgumentParser(description="Compare QB bank ledgers with ERP GL.")
    parser.add_argument("--tb", action="store_true", help="Compare QB trial balance vs ERP closings")
    parser.add_argument("--excel", type=Path, help="QuickBooks Desktop GL Excel export for one bank")
    parser.add_argument("--erp-account", help="ERP account number for --excel compare (e.g. 10008)")
    parser.add_argument(
        "--export-erp",
        type=Path,
        nargs="?",
        const=SCRIPT_DIR / "erp_bank_ledgers",
        help="Export ERP bank ledgers to CSV files",
    )
    args = parser.parse_args()

    if args.export_erp:
        export_erp_ledgers(args.export_erp)
        return 0

    if args.excel:
        if not args.erp_account:
            print("--erp-account is required with --excel", file=sys.stderr)
            return 2
        if not args.excel.exists():
            print(f"File not found: {args.excel}", file=sys.stderr)
            return 2
        return compare_excel_ledger(args.excel, args.erp_account)

    return compare_tb_vs_erp()


if __name__ == "__main__":
    raise SystemExit(main())
