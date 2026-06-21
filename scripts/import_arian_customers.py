"""Import Arian Traders customer list into company 7 (uses Address column, ignores Bill to 2 / City)."""
import re
import sys
from pathlib import Path

import openpyxl
import pyodbc

FILE = Path(r"C:\Users\Muhammad Israr Ali\OneDrive\Desktop\Arian Traders Customers.xlsx")
COMPANY_ID = 7
PREFIX = "BUYER-"
IMPORT_USER = "arian-customer-import"


def normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", (value or "").strip().lower())


def normalize_tax_id(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text or text in {"0", "0.0"}:
        return None
    return text


def main():
    if not FILE.exists():
        print(f"File not found: {FILE}")
        return 1

    wb = openpyxl.load_workbook(FILE, data_only=True)
    ws = wb.active
    headers = {}
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        for col_idx, cell in enumerate(row, start=1):
            key = normalize_header(str(cell) if cell is not None else "")
            if key and key not in headers:
                headers[key] = col_idx
        if "customer" in headers or "customername" in headers:
            header_row = row_idx
            break

    if header_row is None:
        print("Customer column not found.")
        return 1

    name_col = headers.get("customername") or headers.get("customer")
    address_col = headers.get("address")
    ntn_col = headers.get("ntn")
    cnic_col = headers.get("cnic")
    strn_col = headers.get("strn") or headers.get("registration")
    phone_col = headers.get("phone") or headers.get("mobile")
    email_col = headers.get("email")
    province_col = headers.get("province")

    conn = pyodbc.connect(
        "DRIVER={ODBC Driver 17 for SQL Server};SERVER=localhost;DATABASE=PakistanAccountingERP;Trusted_Connection=yes;"
    )
    cur = conn.cursor()

    cur.execute("SELECT Id, UPPER(Name) FROM Provinces WHERE IsActive = 1")
    provinces = {name: pid for pid, name in cur.fetchall()}

    cur.execute(
        "SELECT Id, UPPER(LTRIM(RTRIM(BuyerName))) FROM Customers WHERE CompanyId = ? AND IsDeleted = 0",
        COMPANY_ID,
    )
    existing = {name: cid for cid, name in cur.fetchall()}

    cur.execute(
        "SELECT BuyerId FROM Customers WHERE CompanyId = ? AND BuyerId LIKE ? AND IsDeleted = 0",
        COMPANY_ID,
        PREFIX + "%",
    )
    max_num = 0
    for (buyer_id,) in cur.fetchall():
        suffix = buyer_id[len(PREFIX):]
        if suffix.isdigit():
            max_num = max(max_num, int(suffix))
    next_num = max_num + 1

    imported = updated = skipped = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        name = row[name_col - 1] if name_col and name_col <= len(row) else None
        if name is None or not str(name).strip():
            continue
        name = str(name).strip()
        address = None
        if address_col and address_col <= len(row) and row[address_col - 1]:
            address = str(row[address_col - 1]).strip() or None
        ntn = normalize_tax_id(row[ntn_col - 1] if ntn_col and ntn_col <= len(row) else None)
        cnic = normalize_tax_id(row[cnic_col - 1] if cnic_col and cnic_col <= len(row) else None)
        strn = normalize_tax_id(row[strn_col - 1] if strn_col and strn_col <= len(row) else None)
        phone = row[phone_col - 1] if phone_col and phone_col <= len(row) else None
        phone = str(phone).strip() if phone else None
        email = row[email_col - 1] if email_col and email_col <= len(row) else None
        email = str(email).strip() if email else None
        province_name = row[province_col - 1] if province_col and province_col <= len(row) else None
        province_id = 1
        if province_name:
            key = str(province_name).strip().upper()
            province_id = provinces.get(key, 1)
        customer_type = 1 if ntn else (2 if cnic else 1)
        scenario_id = 1 if ntn else (2 if cnic else 1)

        key = name.upper()
        if key in existing:
            cur.execute(
                """
                UPDATE Customers SET
                    Address = COALESCE(?, Address),
                    Phone = COALESCE(?, Phone),
                    Email = COALESCE(?, Email),
                    NTN = COALESCE(?, NTN),
                    CNIC = COALESCE(?, CNIC),
                    STRN = COALESCE(?, STRN),
                    ProvinceId = ?,
                    CustomerType = ?,
                    ScenarioId = ?,
                    IsActive = 1,
                    UpdatedAt = SYSUTCDATETIME(),
                    UpdatedBy = ?
                WHERE Id = ?
                """,
                address,
                phone,
                email,
                ntn,
                cnic,
                strn,
                province_id,
                customer_type,
                scenario_id,
                IMPORT_USER,
                existing[key],
            )
            updated += 1
            continue

        buyer_id = f"{PREFIX}{next_num:04d}"
        next_num += 1
        cur.execute(
            """
            INSERT INTO Customers (
                CompanyId, BuyerId, BuyerName, OpeningBalance, Address, ProvinceId, ScenarioId,
                Phone, Email, NTN, CNIC, STRN, CustomerType, InvoiceType, IsActive,
                CreatedAt, CreatedBy, IsDeleted)
            VALUES (?, ?, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, 1, SYSUTCDATETIME(), ?, 0)
            """,
            COMPANY_ID,
            buyer_id,
            name,
            address,
            province_id,
            scenario_id,
            phone,
            email,
            ntn,
            cnic,
            strn,
            customer_type,
            IMPORT_USER,
        )
        imported += 1

    conn.commit()
    conn.close()
    print(f"Imported {imported}, updated {updated}, skipped {skipped} for company {COMPANY_ID}.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
