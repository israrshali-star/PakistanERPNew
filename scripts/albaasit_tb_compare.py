"""Compare QuickBooks trial balance xlsx vs ERP SQL export for Al Baasit (company 6)."""
import re
import subprocess
import sys
from pathlib import Path

import openpyxl

QB_TO_ERP = {
    "10020": "10013",
    "10800": "10015",
    "10900": "10016",
    "11000": "11110",
    "12000": "10017",
    "15200": "15100",
    "30800": "30020",
    "10110": "10010",
    "10120": "10012",
    "32000": "32000",
    "25500": "25520",
}

SKIP = {"10000", "12100", "47900"}


def extract_acct(label: str) -> str | None:
    m = re.search(r":(\d{4,5})\b", label) or re.search(r"^\s*\"?(\d{4,5})\b", label)
    return m.group(1) if m else None


def parse_qb(path: Path) -> dict[str, float]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    openings: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        label = str(row[0]).strip()
        if label.lower() in ("debit", "credit") or "jun 30" in label.lower():
            continue
        acct = extract_acct(label)
        if not acct:
            continue
        erp = QB_TO_ERP.get(acct, acct)
        if erp in SKIP:
            continue
        debit = float(row[1] or 0) if len(row) > 1 and row[1] not in (None, "") else 0.0
        credit = float(row[2] or 0) if len(row) > 2 and row[2] not in (None, "") else 0.0
        if debit == 0 and credit == 0:
            continue
        openings[erp] = round(debit - credit, 2)
    wb.close()
    return openings


def parse_erp() -> dict[str, float]:
    q = """
    SELECT AccountNumber, OpeningBalance
    FROM ChartOfAccounts
    WHERE CompanyId=6 AND IsDeleted=0 AND IsActive=1
      AND NOT EXISTS (SELECT 1 FROM ChartOfAccounts p WHERE p.ParentAccountId=ChartOfAccounts.Id AND p.IsDeleted=0)
    """
    out = subprocess.check_output(
        ["sqlcmd", "-S", "localhost", "-d", "PakistanAccountingERP", "-E", "-W", "-h", "-1", "-Q", q],
        text=True,
    )
    erp: dict[str, float] = {}
    for line in out.splitlines():
        parts = line.split()
        if len(parts) < 2:
            continue
        acct, bal = parts[0], parts[1].replace(",", "")
        try:
            erp[acct] = round(float(bal), 2)
        except ValueError:
            pass
    return erp


def main() -> int:
    path = Path(r"C:\Users\Muhammad Israr Ali\OneDrive\Desktop\Al Baasit Trading\Trial Balance as of 306062026.xlsx")
    if "--list-qb" in sys.argv:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            label = str(row[0]).strip()
            acct = extract_acct(label)
            if not acct:
                continue
            d = float(row[1] or 0) if len(row) > 1 and row[1] not in (None, "") else 0.0
            c = float(row[2] or 0) if len(row) > 2 and row[2] not in (None, "") else 0.0
            if d == 0 and c == 0:
                continue
            print(f"{acct}\t{d}\t{c}\t{round(d-c,2)}\t{label[:80]}")
        wb.close()
        return 0

    qb = parse_qb(path)
    erp = parse_erp()
    keys = sorted(set(qb) | set(erp))
    print(f"{'Acct':<8} {'QB':>18} {'ERP':>18} {'Diff':>18}")
    mismatches = 0
    for k in keys:
        qv = qb.get(k, 0.0)
        ev = erp.get(k, 0.0)
        if qv == 0 and ev == 0:
            continue
        diff = round(ev - qv, 2)
        flag = " ***" if diff != 0 else ""
        if diff != 0:
            mismatches += 1
        print(f"{k:<8} {qv:18,.2f} {ev:18,.2f} {diff:18,.2f}{flag}")
    print(f"\nMismatches: {mismatches}")
    print(f"QB sum: {sum(qb.values()):,.2f}  ERP sum: {sum(erp.values()):,.2f}")
    return mismatches


if __name__ == "__main__":
    raise SystemExit(main())
